import customtkinter as ctk
from customtkinter import filedialog 
import time 
from PIL import Image
from CTkTable import *
import pytesseract
import re
import openpyxl 
import datetime
import pandas as pd 
import os
import threading
import playsound
from tkinter import messagebox
from CTkMessagebox import CTkMessagebox
import pyautogui
import logging

def extract_emails(text1):
    if pd.isnull(text1) or not isinstance(text1, str):
        return []
    email_regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    return re.findall(email_regex, text1)

def extract_URL(text1):
    if pd.isnull(text1) or not isinstance(text1, str):
        return []
    URL_regex = r"(?:(?:https?|ftp|file):\/\/|www\.|ftp\.)(?:\([-a-zA-Z0-9+&@#\/%=~_|$?!:,.]*\)|[-a-zA-Z0-9+&@#\/%=~_|$?!:,.])*(?:\([-a-zA-Z0-9+&@#\/%=~_|$?!:,.]*\)|[a-zA-Z0-9+&@#\/%=~_|$])"
    return re.findall(URL_regex, text1)

def extract_phone(text1):
    if pd.isnull(text1) or not isinstance(text1, str):
        return []
    phone_regex = r"[+]?[0-9]{2}\s\d{5}\s\d{5}|[6-9]{1}\d{9}|[0-9]{3}\-[0-9]{8}|[+]?[0-9]{2}\s\d{3}\s?\d{7}|\(?\d{3}\)?\s?\-?\s?\d{4}\s?\d{4}|\d{4}\s\d{7}|\d{7}|[+]\d{2}\s?\-?\s?\d{10}|[+]\d{2}\s?\-?\s?\d{2}\s?\d{4}\s?\d{4}\s?\W?\s?[0-9]*|[+]?\d{2}\s?\d{3}\s?\d{3}\s?\d{4}|[+]?[0-9]{2}\s\d{9}|[+]?[0-9]{2}\.?\d{5}\s?\d{5}|[+]?[0-9]{2}\s?\d{3}\-?\d{7}|[+]?[0-9]{2}\s?\W?\s?\d{5}\s?\d{5}|[+]?[0-9]{2}\W?\d{2}\s?\d{8}|[+]?[0-9]{2}\s?\-\s?\d{2}\s?\-\s?\d{8}\s?\W\s?[0-9]*|[+]?[0-9]{2}\s?\-\s?\d{2}\-\d{4}\s?\d{4}|[+]?[0-9]{2}\s?\d{2}\s?\d{2}\s?\d{6}|\d{5}\s?\d{5}|\d{2}\s\d{4}\s\d{4}"
    return re.findall(phone_regex, text1)


class DataExtractionApp(ctk.CTk):
    
    def __init__(self):
        super().__init__()
        global Sheet1,wb,excel_name,flag1,cell_value,filename_img,config,Sheet,state1
         
         #logging 
        logging.basicConfig(filename="AppLog.log",level=logging.INFO,format='%(asctime)s - %(levelname)s - %(message)s')
        self.geometry("750x605")
        self.title("Info Extract")
        self.iconbitmap("favicon.ico")
        self.resizable(False, False)
        self.grid_columnconfigure((0,1), weight=1)

        flag = 0

        config = openpyxl.load_workbook('Config_1.xlsx')
        Sheet = config.active
        cell_value = Sheet['B2'].value
        file_name = Sheet['B3'].value

        #Image for search logo

        global img,folder_icon,flag1
        img = Image.open("search (2).png")
        folder_icon = Image.open("folder (1).png")

        flag1 = 2
        filename_img =" "
         
        # for state1 used to edit the rename 
        state1 = "readonly"

        home_directory_path = os.path.expanduser("~")
        desktop_path = os.path.join(home_directory_path, "Desktop")


        current_time = datetime.datetime.now().strftime("%d_%m_%Y__%H_%M_%S")
        
        # file name getting  
        if file_name and cell_value:
            excel_name = cell_value+"\\"+file_name+"_"+current_time+".xlsx" 
            logging.info(f"File Path : {excel_name}")
        elif file_name and cell_value is None:
            excel_name = desktop_path+"\\"+file_name+"_"+current_time+".xlsx" 
            logging.info(f"File Path : {excel_name}")
        elif file_name is None and cell_value is None:
            excel_name = desktop_path+r"\WTS_data_extracter_"+current_time+".xlsx" 
            logging.info(f"File Path : {excel_name}")
        elif file_name is None and cell_value:
            excel_name = cell_value+r"\WTS_data_extracter_"+current_time+".xlsx"
            logging.info(f"File Path : {excel_name}")

        wb = openpyxl.Workbook()
        Sheet1 = wb.active
        Sheet1['A1'] = 'Exctacted Data'
        Sheet1['B1'] = 'Image Name'
        Sheet1['C1'] = 'Email'
        Sheet1['D1'] = 'Url'
        Sheet1['E1'] = 'Phone Number'

        #App Title 
        self.App_header_title = ctk.CTkLabel(self, text="WTS Data Extraction Tool", font=ctk.CTkFont(size=30, weight="bold"))
        self.App_header_title.grid(row=0,column=0,padx=10, pady=(10,0),columnspan=3)
        
        #Switch Mode
        self.Mode_SB = ctk.CTkSegmentedButton(self,values=["Individual Mode","Bulk Mode","Settings"],corner_radius=30,border_width=5,font=ctk.CTkFont(size=15),command=self.Mode_switch)
        self.Mode_SB.grid(row=1, column=0, pady=(15,0),columnspan=3)
        self.Mode_SB.set("Individual Mode") #set initial value
        self.Mode_switch("Individual Mode") #call the the mode with initial value

        self.copyright_lable = ctk.CTkLabel(self,text="© 2024 WiseTechSource Pvt. Ltd. All rights reserved.",font=ctk.CTkFont(family="times",size=15))
        self.copyright_lable.grid(row=4,column=0,padx=(30,0),pady=(0,20))

        self.protocol("WM_DELETE_WINDOW", self.close_window)

        #Get the mode 

    def Mode_switch(self,value):
            global header_list
            if hasattr(self, 'frame2'):
                self.frame2.destroy()  
            if value == "Individual Mode":
                mode_text = "Image Mode , select the image path"
                R1_text = "Input Image Path"
                status_text =" "

                header_list = [["Email","URL","Phone Number"],]

                browse_img = self.get_image_path
                submit_img = self.extract_button_img

                self.frame(mode_text,R1_text,status_text,browse_img,submit_img)

                self.M1_Scroll =ctk.CTkScrollableFrame(self.frame1,width=100,height=50,corner_radius=15,border_width=2,border_color="gray",orientation="vertical")
                self.M1_Scroll.grid(row=3, column=0, columnspan=3, padx=(30,10),pady=(0,10), sticky="ew")

                self.table = CTkTable(self.M1_Scroll,column=3,values=header_list,width=192,font=ctk.CTkFont("times",15,"bold"))
                self.table.grid(row=0, column=0, sticky="nsew")
            
            elif value == "Bulk Mode":

                mode_text = "Multiple files Mode , select the folder path of images"
                R1_text = "Input Folder Path"
                status_text=" "

                Folder_path = self.get_folder_path
                submit_folder = self.folder_extracter


                self.frame(mode_text,R1_text,status_text,Folder_path,submit_folder)

                self.count_text_lab = ctk.CTkLabel(self.frame1,text="Number of Images in the folder :",font=ctk.CTkFont(family="times",size=17))
                self.count_text_lab.grid(row = 4, column =0 ,pady = 10 , padx = (100,0),columnspan=2)

                self.count_lab = ctk.CTkLabel(self.frame1,text="0",font=ctk.CTkFont(family="times",size=17))
                self.count_lab.grid(row = 4, column =1 ,pady = 10,padx=(250,0)) 
                
            elif value == "Settings":

                S1_text = "Destination Folder Path"
                S2_text = "Enter the File Name"
                self.setting(S1_text,S2_text)

    def frame(self,mode_text,R1_text,status_text,browse_type,sumbit_type):
         
         self.frame1=ctk.CTkFrame(self,corner_radius=15,height=400,width=630)                      
         self.frame1.grid(row=2, column=0, padx=(26,0), pady=10,sticky="news")    

         self.mode_lab1=ctk.CTkLabel(self.frame1,text=mode_text,font=ctk.CTkFont(family="Arial",size=15),text_color="#f7e09c" )
         self.mode_lab1.grid(row=1,column=0,padx=(0,0),pady=(15,5),columnspan=3) 

         self.R1_text_lab=ctk.CTkLabel(self.frame1,text=R1_text,font=ctk.CTkFont(family="times",size=20,weight="bold")) 
         self.R1_text_lab.grid(row=2,column=0,padx=(20,0))

         self.R2_text_entry =ctk.CTkEntry(self.frame1,width=330,placeholder_text="Click browse button..",corner_radius=30,border_color="#3a7ea6",height=35)
         self.R2_text_entry.grid(row=2,column=1,padx=(20,0))

         self.R3_browse_button = ctk.CTkButton(self.frame1,text="Browse",font=ctk.CTkFont("family",15,'bold'),command=browse_type, image=ctk.CTkImage(dark_image=img,light_image=img),fg_color="#454746",corner_radius=30,height=35,compound="right")
         self.R3_browse_button.grid(row=2,column=2,padx=(15,0),pady=(20))

         self.Status_lab = ctk.CTkLabel(self.frame1,text=status_text)
         self.Status_lab.grid(row=5,column=1,columnspan=3,pady=(10,0),padx=(0,100))

         self.sumbit_buttom = ctk.CTkButton(self.frame1,text="Extract",corner_radius=30,font=ctk.CTkFont(family="Arial",size=15,weight="bold"),height=35,command=sumbit_type)
         self.sumbit_buttom.grid(row=6,column=1,padx=(40,0),pady=(5,10))
        
    def setting(self,s1_text=None,S2_text=None):
         
        

         self.frame2=ctk.CTkFrame(self,corner_radius=15,height=400,width=630)
         self.frame2.grid(row=2, column=0, padx=(26,0), pady=10,sticky="news")

         self.S1_Path_lab =ctk.CTkLabel(self.frame2,text=s1_text,font=ctk.CTkFont(family="times",size=15),text_color="green")
         self.S1_Path_lab.grid(row=1,column=0,pady=(20,0),padx=(10,20))

         self.S2_Destination_entry = ctk.CTkEntry(self.frame2,width=400,corner_radius=15,height=35,border_color="green",placeholder_text="Select the path of destination.. ")
         self.S2_Destination_entry.grid(row =2, column =0 ,pady=10,padx=(30,30))
         
         
         config = openpyxl.load_workbook('Config_1.xlsx')
         Sheet = config.active
         cell_value = Sheet['B2'].value
         File_name = Sheet['B3'].value

         if cell_value:
            self.S2_Destination_entry.delete(0,ctk.END)
            self.S2_Destination_entry.insert(0,cell_value)
            self.S2_Destination_entry.configure(state="readonly")
        
         self.button_out_path = ctk.CTkButton(self.frame2,text="Browse",font=ctk.CTkFont(size=15),corner_radius=30,image=ctk.CTkImage(dark_image=folder_icon,light_image=folder_icon),fg_color="#454746",height=35,width=5,command=self.output_path)
         self.button_out_path.grid(row=2,column=1,padx=(0,0),pady=(10))

         self.S1_Path_lab =ctk.CTkLabel(self.frame2,text=S2_text,font=ctk.CTkFont(family="times",size=15),text_color="green")
         self.S1_Path_lab.grid(row=3,column=0,pady=(20,0),padx=(10,40))

         self.S2_Rename_entry = ctk.CTkEntry(self.frame2,width=400,corner_radius=15,height=35,border_color="green",placeholder_text="Enter the file name..",)
         self.S2_Rename_entry.grid(row =4, column =0 ,pady=10,padx=(0,0))

         if File_name:
             self.S2_Rename_entry.insert(0,File_name)

         self.S2_Rename_entry.configure(state=state1)
         self.S2_Rename_entry.bind('<Key>', self.validate_filename_entry)

         self.button_out_path = ctk.CTkButton(self.frame2,text="Sumbit",font=ctk.CTkFont(size=15),corner_radius=30,fg_color="#454746",height=35,width=50,command=self.rename_submit)
         self.button_out_path.grid(row=4,column=1,padx=(0,0),pady=(10))

         self.button_out_path = ctk.CTkButton(self.frame2,text="Edit",font=ctk.CTkFont(size=15),corner_radius=30,fg_color="#454746",height=35,width=80,command=self.edit_button)
         self.button_out_path.grid(row=4,column=2,padx=(0,0),pady=(10))

    def get_image_path(self):
      global filename_img
      filename_img = ""
      filename_img = filedialog.askopenfilename(initialdir="/", title="Select Image",filetypes=(("Image file", "*.jpg; *.jpeg; *.png;"), ))
      if filename_img:
            self.mode_lab1.configure(text="Path selected succesfully",text_color="#f7e09c")
            print("Selected Image Path:", filename_img)
            self.R2_text_entry.delete(0,ctk.END)
            self.R2_text_entry.insert(0,filename_img)  
            
    def extract_button_img(self):
        global flag1,filename_img
        if hasattr(self, 'R2_text_entry'):

            Pure_file_path = self.R2_text_entry.get()
            
            if (self.R2_text_entry.get() and os.path.exists(Pure_file_path)) :
                self.Individual_Mode(Pure_file_path)
                
            elif self.R2_text_entry.get() == "":
                self.mode_lab1.configure(text="Please Select Any File Path",text_color ="red")
            else:
                self.mode_lab1.configure(text="Please Select The Valid Image Path ",text_color ="red")  
  
    def get_folder_path(self):
        
        global filename_folder,img_count

        filename_folder = filedialog.askdirectory(initialdir="/", title="Select Folder")
        self.R2_text_entry.delete(0,ctk.END)
        self.R2_text_entry.insert(0,filename_folder)
        image_extensions = ['.jpg', '.jpeg', '.png']  # Add more if needed
        img_count = 0
        files = os.listdir(filename_folder)
        for file in files:
            _, ext = os.path.splitext(file)
            if ext.lower() in image_extensions:
                img_count += 1
        self.count_lab.configure(text=img_count) 
        self.mode_lab1.configure(text="")
        self.mode_lab1.configure(text="Path Selected Successfully")
        
    def output_path(self):
        
        global folder_path_out
        self.S2_Destination_entry.configure(state="normal")
        folder_path_out = filedialog.askdirectory(initialdir="/", title="Select Folder")
        if folder_path_out:
            self.S2_Destination_entry.delete(0,ctk.END)
            self.S2_Destination_entry.insert(0,folder_path_out)
            Sheet['B2'] = folder_path_out
            config.save('Config_1.xlsx')
            config.close()
            self.S2_Destination_entry.configure(state="readonly")
            CTkMessagebox(title="Updated",message="Successfully Path Updated",icon="check", option_1="OK")
        else:
            folder_path_out = ""
            CTkMessagebox(title="Not Updated",message="Not Updated",icon="warning", option_1="OK")

    def folder_extracter(self):
        if self.R2_text_entry.get() and os.path.exists(self.R2_text_entry.get().replace('"','')):
                t = threading.Thread(target=self.bulk_method, args=(self.R2_text_entry.get(),))
                t.start()  
        elif self.R2_text_entry.get()=="":
            self.Status_lab.configure(text="Please Select Any Folder Path",text_color="red",font = ctk.CTkFont(size=15))     
        else:
          self.Status_lab.configure(text="Please Select Correct Folder Path",text_color="red",font = ctk.CTkFont(size=15))     

    def bulk_method(self,folder_path):
        global flag
        if cell_value:
            if os.path.exists(cell_value):
                try:

                    self.Status_lab.configure(text=" ",text_color="White")     
                    self.Mode_SB.configure(state="disabled")
                    self.R3_browse_button.configure(state="disabled")
                    self.sumbit_buttom.configure(state="disabled")

                    # Set the path to the Tesseract executable (change it according to your installation)
                    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

                    list_of_file = os.listdir(folder_path)
                    image_extensions = ['.jpg', '.jpeg', '.png']
                    flag = 2

                    for current_image in list_of_file:
                        base_name , extention = os.path.splitext(current_image)

                        if extention in image_extensions:
                            try:

                                #tesseract ocr --------------------------------------------------------------------
                                # Open an image using PIL
                                image_path = folder_path+'/'+current_image
                                img = Image.open(image_path)
                                
                                # Use Tesseract to do OCR on the image
                                text1 = pytesseract.image_to_string(img)
                                
                                curr_img = flag-1
                                progress_value = ((curr_img) / img_count )
                                
                                self.loading_bar = ctk.CTkProgressBar(self.frame1,orientation="horizontal",height=25,width=300,border_width=2,progress_color="#5b86b5",border_color="#fcfcfc")
                                self.loading_bar.grid(row=7,column =1,pady=25,padx = (20,0))
                                self.loading_bar.set(progress_value)
                                loading_bar_var = "("+str(curr_img)+"/"+str(img_count)+")"

                                

                                self.loading_text_lab = ctk.CTkLabel(self.frame1,text="",font=ctk.CTkFont(family="times",size=17))
                                self.loading_text_lab.grid(row=8,column=1)
                                self.loading_text_lab.configure(text="")
                               
                                self.loading_text_lab.configure(text=loading_bar_var)

                                
                                #Print the extracted text
                                print("Extracted Text:",text1)
                                #tesseract ocr---------------------------------------------------------------------
                                
                                email = extract_emails(text1)
                                URL = extract_URL(text1)
                                phone = extract_phone(text1)

                                if email:
                                    email =email
                                else:
                                    email = "-"
                                if URL:
                                    URL =URL
                                else:
                                    URL = "-"
                                if phone:
                                    phone =phone   
                                else:
                                    phone = "-"
                                
                                Sheet1['A'+str(flag)] = text1.replace("\n\n","\n")
                                Sheet1['B'+str(flag)] = current_image
                                Sheet1['C'+str(flag)] = " | ".join(email) if isinstance(email, tuple) else str(" ".join(email))
                                Sheet1['D'+str(flag)] = " | ".join(URL) if isinstance(URL, tuple) else str(" ".join(URL))
                                Sheet1['E'+str(flag)] = " | ".join(phone) if isinstance(phone, tuple) else str(phone)

                                logging.info(f"Mode 2 : Successfully Extracted - Image File Name {current_image}")

                                flag=flag+1

                                #save the excel file
                                wb.save(excel_name)

                            except Exception as err:
                                logging.error(f"Mode 2 ; Unable To Identify The Image Name : {current_image}")    
                    wb.close()

                    self.Mode_SB.configure(state="active")
                    self.R3_browse_button.configure(state="normal")
                    self.sumbit_buttom.configure(state="normal")

                    CTkMessagebox(title="Successfully Extracted",message="Hooray! Successfully Completed ",icon="check")
                    

                except Exception as err:
                    print(err)
                    
                               
        else:
            CTkMessagebox("Error","Please select folder path , Check the settings")
        self.R2_text_entry.delete(0,ctk.END)         
    
    def Individual_Mode(self,File_type):  
        try:
            global flag1
            self.Mode_SB.configure(state="disabled")
            #tesseract ocr -------------------

            # Open an image using PIL
            img = Image.open(File_type)
            # Use Tesseract to do OCR on the image
            text1 = pytesseract.image_to_string(img)
            print("Extracted Text:",text1)

            #tesseract ocr---------------------
            email = extract_emails(text1)
            URL =extract_URL(text1)
            Phone =extract_phone(text1)

            if email:
                email =email
            else:
                email = "-"
            if URL:
                URL =URL
            else:
                URL = "-"  
            if Phone:
                Phone =Phone
            else:
                Phone = "-"   

            Sheet1['A'+str(flag1)] = text1.replace("\n\n","\n")
            Sheet1['B'+str(flag1)] = os.path.basename(File_type)
            Sheet1['C'+str(flag1)] = " | ".join(email) if isinstance(email, tuple) else str(" ".join(email))
            Sheet1['D'+str(flag1)] = " | ".join(URL) if isinstance(URL, tuple) else str(" ".join(URL))
            Sheet1['E'+str(flag1)] = " | ".join(Phone) if isinstance(Phone, tuple) else str(Phone)
            logging.info(f"Mode 1 ;  Successfully Extracted - Image File Name  : {os.path.basename(File_type)}")
            flag1 = flag1 +1

            #save the excel file
            
            wb.save(excel_name)
        
            newlist =[email,URL,Phone]
            Exceldata_list = [text1,email,URL,Phone,File_type]
            header_list.append(newlist)
            table = CTkTable(self.M1_Scroll, column=3, values=header_list, width=183,font=ctk.CTkFont("times",15))
            table.grid(row=0, column=0, sticky="nsew")
            self.R2_text_entry .delete(0,ctk.END)
            self.Mode_SB.configure(state="normal")
            self.Status_lab.configure(text =" ")
            self.mode_lab1.configure(text="Image Mode , select the image path",text_color ="#f7e09c")

        except Exception as err:
            print(str(err)+" error")
            self.Mode_SB.configure(state="normal")
            self.mode_lab1.configure(text="Unable to identify the image file",text_color="#b50e0e")
            logging.error(f"Mode 1 ; Unable To Identify The Image Name : {os.path.basename(File_type)}") 
 
    def validate_filename_entry(self, event):
       
        current_text = self.S2_Rename_entry.get()

        # Regular expression pattern to match special characters
        pattern = r'[\\/:*?<>|]'

        # Check if the entered text contains any special characters
        if re.search(pattern, current_text):
            # Special character found, delete the last character entered
            self.S2_Rename_entry.delete(len(current_text) - 1, ctk.END)
            # Display a message to the user about the restriction (optional)
            CTkMessagebox(title="Error",message="Special characters are not allowed in the filename.",icon="warning")
            
    def rename_submit(self):
        global state1
        pyautogui.press("enter")
        state1 = 'readonly'
        pattern = r'[\\/:*?<>|]' 
        if re.search(pattern,self.S2_Rename_entry.get()):
            print("true")
        else:
            print(self.S2_Rename_entry.get())
            Sheet['B3'] = self.S2_Rename_entry.get()
            config.save('Config_1.xlsx')
            config.close()
            logging.info(f"File Name Renamed : {self.S2_Rename_entry.get()}")
            self.setting("Destination Folder Path","Enter the File Name")

    def edit_button(self):
        global state1
        state1 = "normal"
        self.setting("Destination Folder Path","Enter the File Name")

    def close_window(self):
        logging.info("Execution Stoped.")
        self.destroy()

   
app = DataExtractionApp()
app.mainloop()

       